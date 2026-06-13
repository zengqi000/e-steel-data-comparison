"""
销售合同 vs 销售实提 重量核对脚本
对比完成后直接输出 Excel 到 stdout（base64），供 WorkBuddy 弹窗下载
用法：
  python compare_contracts.py "合同名1" "合同名2" ...   指定合同名
  python compare_contracts.py --date-range 20260610-20260612  按时间段查询
  python compare_contracts.py                           查询全部合同
"""

import sys
import re
import requests
import json
import base64
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# 配置区域 - 只需修改 TOKEN
# ============================================================

TOKEN = 'Bearer eyJhbGciOiJSUzI1NiIsInR5cCI6IkpXVCJ9.eyJwa0lkIjoiMTYwMzAyMDU4MTgzMTEzOTMyOSIsInVzZXJfbmFtZSI6IiZBRE1JTiZjZW5ncWlAaXRnLm5ldCIsIm9wZW5JZCI6IjYyMDBjNjI2ZTRiMDYwZjU5ZWMzN2NjMiIsIm9hQ29tcGFueU5hbWUiOiLljqbpl6jlm73otLjmlbDlrZfnp5HmioDmnInpmZDlhazlj7giLCJjbGllbnRfaWQiOiJpdGctZ2F0ZXdheSIsIndvcmtOdW0iOiI1MDQ4NzUiLCJzYXBEZXB0Q29kZSI6IlNLNiIsInNhcERlcHRJZCI6MjAsInNjb3BlIjpbImFsbCJdLCJjb250ZXh0VHlwZSI6ImFkbWluX2NvbnRleHQiLCJjb21wYW55Ijoi5Y6m6Zeo5Zu96LS45pWw5a2X56eR5oqA5pyJ6ZmQ5YWs5Y-4Iiwic2FwQ29tcGFueUNvZGUiOiJDT00wMDAwMjU1IiwiZXhwIjoxNzgxMzM0Nzk3LCJqdGkiOiI0ZjdhODdhOC0yNzk4LTRkZjYtOTFmZC00ZTRkZjFhYWRmMTYiLCJvYURlcHRJZCI6IjEzNDMiLCJjb21wYW55Q29kZSI6IkNPTTAwMDAyNTUiLCJzYXBDb21wYW55SWQiOjcxLCJ0ZWxlcGhvbmUiOiIxODcyMDkzNjA2MSIsImRlcHQiOiLmtYvor5Xnu4QiLCJvYURlcHROYW1lIjoi6ZKi6ZOB5Lia5Yqh5pWw5a2X5YyW6YOoIiwicGhvbmUiOiIxODcyMDkzNjA2MSIsIm5hbWUiOiLmm77nkKoiLCJkZXB0Q29kZSI6IjMwMzQ2MiIsIm9hSWQiOiIxNDcxNyIsInNhcENvbXBhbnlOYW1lIjoi5Y6m6Zeo5Zu96LS45pWw5a2X56eR5oqA5pyJ6ZmQ5YWs5Y-4Iiwic2FwRGVwdE5hbWUiOiLCoOWbvei0uOaVsOenkemDqOmHkeWxnuS4muWKoeaUr-aMgemDqCIsInVzZXJuYW1lIjoiY2VuZ3FpQGl0Zy5uZXQiLCJvYUNvbXBhbnlJZCI6IjIzMSJ9.HcRs06h5ezAiSyJX7cn3U0fSFpt7AbDN8K4ZSjOMJdocYxUyTalxfYpnYmktMHp6PGQMJj8m8atVBFbVQMij48-CuJxuAYVE1PFQkNgzSjK6-eiy3g_uSxrJugFq9zAOkKQSzwab0Oyd7jrW1hcvD07elp-M-z4m0i5fTPE2DipNpejsPS2hm2hPw_nqIzDdcdkot6-zyqZrAnpvsozBcogvJhrR-8PKu4DtItVHqR6LPTITYAdzHZUVmUthiBIyqLyVUcoIPSUCH8y1CUahOglbdRwdc5pN0xqDlaWP-rGZ834n_66ofhSpFQuRsTrSst7QRCYazfQk2XMg0oNNVg'

HEADERS = {
    "accept": "application/json",
    "authorization": TOKEN,
    "content-type": "application/json;charset=UTF-8",
}

WEIGHT_TOLERANCE = 0.001

# ============================================================
# 解析合同号
# ============================================================

def _is_token_expired(msg: str) -> bool:
    """检测接口返回的消息是否表示 TOKEN 过期"""
    keywords = ["token过期", "token已过期", "token expired", "登录过期", "登录已过期",
                "登录失效", "认证失败", "身份验证", "授权过期", "凭证过期", "凭证失效",
                "访问令牌", "access token", "未授权", "unauthorized", "非法令牌",
                "令牌过期", "令牌失效", "会话过期", "会话失效", "session expired"]
    msg_lower = msg.lower()
    return any(kw in msg_lower for kw in keywords)


def get_contract_list() -> list:
    """从命令行参数解析合同列表，支持三种模式：
    1. --date-range 20260610-20260612  按时间段查询合同
    2. "合同名1" "合同名2" ...         指定合同名
    3. 无参数                          查询全部合同
    """
    args = [a.strip() for a in sys.argv[1:] if a.strip()]

    # 模式1：按时间段查询
    if args and args[0] == "--date-range":
        if len(args) < 2:
            print("❌ 请提供时间段，格式：--date-range 20260610-20260612", file=sys.stderr)
            sys.exit(1)
        return fetch_contract_list_by_date(args[1])

    # 模式2：指定合同名
    if args:
        return args

    # 模式3：查询全部
    return []


# ============================================================
# 按时间段查询销售合同列表
# ============================================================

def parse_date_range(date_range_str: str):
    """解析时间段字符串，返回 (start_ts_ms, end_ts_ms)
    支持格式：20260610-20260612 或 20260610
    """
    match = re.match(r'^(\d{8})(?:-(\d{8}))?$', date_range_str)
    if not match:
        raise ValueError(f"时间段格式错误：{date_range_str}，正确格式：20260610-20260612")

    start_str = match.group(1)
    end_str = match.group(2) or start_str  # 如果只传一个日期，开始=结束

    start_dt = datetime.strptime(start_str, "%Y%m%d")
    end_dt = datetime.strptime(end_str, "%Y%m%d")

    # 结束日期取当天 23:59:59.999
    end_dt = end_dt.replace(hour=23, minute=59, second=59, microsecond=999000)

    start_ms = int(start_dt.timestamp() * 1000)
    end_ms = int(end_dt.timestamp() * 1000)

    return start_ms, end_ms


def fetch_contract_list_by_date(date_range_str: str) -> list:
    """根据时间段查询销售合同列表，返回 outContractNo 列表"""
    start_ms, end_ms = parse_date_range(date_range_str)

    url = "https://manage.itgmetals.com/api/itg-es-scroll-webapp/sell-contract/scroll"
    body = {
        "contractNoList": [],
        "outContractNoList": [],
        "sapCodeList": [],
        "orgIdList": [],
        "deptGroupIdList": [],
        "customerIdList": [],
        "budgetContractNoList": [],
        "saleManIdList": [],
        "merchandiserId": [],
        "createAdList": [],
        "sellContractTypeList": [],
        "queryStatus": 9,
        "contractDateStart": start_ms,
        "contractDateEnd": end_ms
    }

    all_records = []
    # 分页拉取所有数据
    page = 1
    while True:
        body["pageNo"] = page
        body["pageSize"] = 200
        resp = requests.post(url, headers=HEADERS, json=body, timeout=30)
        resp.raise_for_status()
        result = resp.json()
        if result.get("code") != 200:
            msg = result.get("msg", "")
            if _is_token_expired(msg):
                print("__TOKEN_EXPIRED__", file=sys.stderr)
                sys.exit(99)
            raise ValueError(f"销售合同列表接口异常: {msg}")

        data = result.get("data") or []
        if not data:
            break
        all_records.extend(data)
        # 如果返回数据不足一页，说明没有更多了
        if len(data) < 200:
            break
        page += 1

    # 提取 outContractNo 并去重
    contract_names = list(dict.fromkeys(
        str(r.get("outContractNo", "")).strip()
        for r in all_records
        if str(r.get("outContractNo", "")).strip()
    ))

    start_display = datetime.fromtimestamp(start_ms / 1000).strftime("%Y-%m-%d")
    end_display = datetime.fromtimestamp(end_ms / 1000).strftime("%Y-%m-%d")
    print(f"📅 查询时间段：{start_display} ~ {end_display}")
    print(f"📋 查到 {len(contract_names)} 个销售合同")

    return contract_names

# ============================================================
# API 请求
# ============================================================

def fetch_contract(contract_list: list) -> list:
    url = "https://manage.itgmetals.com/api/itg-es-scroll-webapp/sell-contract-product-doc/sell-contract-product-check-scroll"
    body = {
        "contractNoList": [], "outContractNoList": contract_list,
        "sapCodeList": [], "customerIdList": [], "brandIdList": [],
        "orgIdList": [], "deptGroupIdList": [], "saleManIdList": [],
        "merchandiserIdList": [], "createAdList": [], "isGoodsCompleteList": [],
        "isInvoiceCompleteList": [], "isMoneyCompleteList": [],
        "istBusinessTagsList": [], "istBusinessSegmentationList": [],
        "finalPriceStatusList": [], "isSettlePriceFinalApproveList": [],
        "dataScop": None, "status": None
    }
    resp = requests.post(url, headers=HEADERS, json=body, timeout=30)
    resp.raise_for_status()
    result = resp.json()
    if result.get("code") != 200:
        msg = result.get("msg", "")
        if _is_token_expired(msg):
            print("__TOKEN_EXPIRED__", file=sys.stderr)
            sys.exit(99)
        raise ValueError(f"接口1异常: {msg}")
    return result.get("data") or []


def fetch_pickup(contract_list: list) -> list:
    url = "https://manage.itgmetals.com/api/itg-es-scroll-webapp/sell-real-pickup-product/scroll-check"
    body = {
        "billNoList": [], "sapCodeList": [], "realPickupTypeList": [],
        "pickupNoList": [], "sellContractNoList": [],
        "sellOutContractNoList": contract_list,
        "sellContractSapNoList": [], "settleCustomerIdList": [], "warehouseIdList": [],
        "orgIdList": [], "deptGroupIdList": [], "purchaseContractNoList": [],
        "purchaseOutContractNoList": [], "purchaseContractSapNoList": [],
        "brandIdList": [], "placeIdList": [], "materialList": [], "specList": [],
        "instoreBillNoAndProductNos": [], "sapBatchNoList": [], "baleNoList": [],
        "productCodeList": [], "saleManIdList": [], "merchandiserIdList": [],
        "createAdList": [], "dataScop": 0
    }
    resp = requests.post(url, headers=HEADERS, json=body, timeout=30)
    resp.raise_for_status()
    result = resp.json()
    if result.get("code") != 200:
        msg = result.get("msg", "")
        if _is_token_expired(msg):
            print("__TOKEN_EXPIRED__", file=sys.stderr)
            sys.exit(99)
        raise ValueError(f"接口2异常: {msg}")
    return result.get("data") or []

# ============================================================
# 数据处理 & 对比
# ============================================================

def build_contract_map(records):
    result = {}
    for r in records:
        key = str(r.get("outContractNo", "")).strip()
        if key:
            result[key] = result.get(key, 0) + float(r.get("sumRealPickWeight") or 0)
    return result


def build_pickup_map(records):
    result = {}
    for r in records:
        key = str(r.get("sellOutContractNo", "")).strip()
        if key:
            result[key] = result.get(key, 0) + float(r.get("weight") or 0)
    return result


def compare(contract_map, pickup_map):
    diff_rows, only_contract, only_pickup = [], [], []
    for key in sorted(set(contract_map) | set(pickup_map)):
        in_c, in_p = key in contract_map, key in pickup_map
        if in_c and in_p:
            wc, wp = contract_map[key], pickup_map[key]
            diff = round(wc - wp, 3)
            if abs(diff) > WEIGHT_TOLERANCE:
                diff_rows.append({
                    "合同名称": key,
                    "合同表-实提重量(吨)": round(wc, 3),
                    "实提表-重量汇总(吨)": round(wp, 3),
                    "差异(吨)": diff,
                })
        elif in_c:
            only_contract.append({"合同名称": key, "合同表-实提重量(吨)": round(contract_map[key], 3)})
        else:
            only_pickup.append({"合同名称": key, "实提表-重量汇总(吨)": round(pickup_map[key], 3)})
    return diff_rows, only_contract, only_pickup

# ============================================================
# 生成 Excel（写入内存，不落盘）
# ============================================================

def build_excel(diff_rows, only_contract, only_pickup, contract_map, pickup_map, contract_list) -> bytes:
    wb = Workbook()
    RED   = PatternFill("solid", start_color="FFCCCC")
    YEL   = PatternFill("solid", start_color="FFF2CC")
    GREEN = PatternFill("solid", start_color="E2EFDA")
    HDR_F = PatternFill("solid", start_color="1F4E79")
    H_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    T_FONT = Font(bold=True, name="Arial", size=12)
    N_FONT = Font(name="Arial", size=10)
    CTR = Alignment(horizontal="center", vertical="center")
    LFT = Alignment(horizontal="left", vertical="center")
    BD  = Border(left=Side(style="thin"), right=Side(style="thin"),
                 top=Side(style="thin"), bottom=Side(style="thin"))

    def write_sheet(ws, title, cols, rows, fill_fn=None):
        ws.merge_cells(f"A1:{get_column_letter(len(cols))}1")
        ws["A1"] = title
        ws["A1"].font = T_FONT
        ws["A1"].alignment = CTR
        ws.row_dimensions[1].height = 26
        ws.append(cols)
        for ci in range(1, len(cols) + 1):
            c = ws.cell(row=2, column=ci)
            c.font = H_FONT; c.fill = HDR_F; c.alignment = CTR; c.border = BD
        ws.row_dimensions[2].height = 20
        for row in rows:
            ws.append(list(row.values()))
            ri = ws.max_row
            fill = fill_fn(row) if fill_fn else None
            for ci in range(1, len(cols) + 1):
                c = ws.cell(row=ri, column=ci)
                c.font = N_FONT; c.border = BD; c.alignment = LFT
                if fill:
                    c.fill = fill
        for ci, col in enumerate(cols, 1):
            ws.column_dimensions[get_column_letter(ci)].width = max(18, len(col) + 4)

    # 汇总
    ws0 = wb.active
    ws0.title = "📊 汇总概览"
    query_str = "、".join(contract_list) if contract_list else "全部合同"
    summary = [
        ["核对时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["查询合同", query_str],
        ["合同表合同数", len(contract_map)],
        ["实提表合同数（去重）", len(pickup_map)],
        ["两表均有合同数", len(set(contract_map) & set(pickup_map))],
        ["重量差异条数", len(diff_rows)],
        ["仅合同表有", len(only_contract)],
        ["仅实提表有", len(only_pickup)],
        ["核对结论", "✅ 无差异" if not diff_rows else f"❌ 发现 {len(diff_rows)} 条重量差异"],
    ]
    ws0.column_dimensions["A"].width = 22
    ws0.column_dimensions["B"].width = 50
    for row in summary:
        ws0.append(row)
        ri = ws0.max_row
        ws0.cell(ri, 1).font = Font(bold=True, name="Arial", size=10)
        c2 = ws0.cell(ri, 2)
        c2.font = Font(name="Arial", size=10)
        if row[0] == "核对结论":
            c2.fill = GREEN if not diff_rows else RED
            c2.font = Font(bold=True, name="Arial", size=10)

    # 差异明细
    if diff_rows:
        ws1 = wb.create_sheet("❌ 重量差异明细")
        write_sheet(ws1, f"重量差异明细（共 {len(diff_rows)} 条）",
                    list(diff_rows[0].keys()), diff_rows, fill_fn=lambda r: RED)
    else:
        ws1 = wb.create_sheet("✅ 无差异")
        ws1["A1"] = "✅ 两表实提重量完全一致，无差异！"
        ws1["A1"].font = Font(bold=True, color="375623", name="Arial", size=12)
        ws1["A1"].fill = GREEN

    if only_contract:
        ws2 = wb.create_sheet("⚠️ 仅合同表有")
        write_sheet(ws2, f"仅合同表存在（共 {len(only_contract)} 条）",
                    list(only_contract[0].keys()), only_contract, fill_fn=lambda r: YEL)

    if only_pickup:
        ws3 = wb.create_sheet("⚠️ 仅实提表有")
        write_sheet(ws3, f"仅实提表存在（共 {len(only_pickup)} 条）",
                    list(only_pickup[0].keys()), only_pickup, fill_fn=lambda r: YEL)

    rows5 = [{"合同名称": k, "sumRealPickWeight(吨)": round(v, 3)} for k, v in sorted(contract_map.items())]
    if rows5:
        write_sheet(wb.create_sheet("📋 合同表数据"),
                    "合同表 - 各合同实提重量（sumRealPickWeight）",
                    list(rows5[0].keys()), rows5)

    rows6 = [{"合同名称": k, "weight汇总(吨)": round(v, 3)} for k, v in sorted(pickup_map.items())]
    if rows6:
        write_sheet(wb.create_sheet("📋 实提表汇总"),
                    "实提表 - 各合同实提重量汇总（weight之和）",
                    list(rows6[0].keys()), rows6)

    # 写入内存 buffer，不落盘
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

# ============================================================
# 主流程
# ============================================================

def main():
    contract_list = get_contract_list()

    contract_records = fetch_contract(contract_list)
    pickup_records   = fetch_pickup(contract_list)

    contract_map = build_contract_map(contract_records)
    pickup_map   = build_pickup_map(pickup_records)

    diff_rows, only_contract, only_pickup = compare(contract_map, pickup_map)

    # 输出文字摘要到 stdout（供 WorkBuddy 读取）
    summary_lines = [
        "📊 核对完成",
        f"- 合同表合同数：{len(contract_map)} 个",
        f"- 实提表合同数：{len(pickup_map)} 个",
        f"- 重量差异条数：{len(diff_rows)} 条",
        f"- 仅合同表有：{len(only_contract)} 条",
        f"- 仅实提表有：{len(only_pickup)} 条",
    ]
    if diff_rows:
        summary_lines.append("")
        summary_lines.append("差异明细：")
        for r in diff_rows:
            summary_lines.append(f"  [{r['合同名称']}]")
            summary_lines.append(f"    合同表={r['合同表-实提重量(吨)']}吨  实提表={r['实提表-重量汇总(吨)']}吨  差异={r['差异(吨)']}吨")
    else:
        summary_lines.append("✅ 所有合同重量核对一致，无差异！")

    print("\n".join(summary_lines))

    # 输出 Excel 为 base64，供 WorkBuddy 弹窗下载
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"重量核对报告_{ts}.xlsx"
    excel_bytes = build_excel(diff_rows, only_contract, only_pickup, contract_map, pickup_map, contract_list)
    b64 = base64.b64encode(excel_bytes).decode()

    # WorkBuddy 识别的下载标记格式
    print(f"\n__WORKBUDDY_FILE_START__")
    print(f"filename={filename}")
    print(f"mimetype=application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    print(f"data={b64}")
    print(f"__WORKBUDDY_FILE_END__")


if __name__ == "__main__":
    main()